# Excel Merger
# By Professor Colin Turner <c.turner@ulster.ac.uk>
# GPL v3.0

"""
This script is intended to map several submitted survey submissions (which are Excel files) into a master document.

It uses a mapping configuration file to work out what cells should go where. It is extremely simple and has almost
no error checking at this time, but I needed something quick and dirty for a task.

It has been used to merge real data however, so feel free to use away or adapt it.

It required OpenPyXL, shout out to that team!
"""

import argparse
import os
import re
import sys

import openpyxl


def override_arguments(args):
    """If necessary, prompt for (what are normally command line) arguments and override them

    Takes, as input, args from an ArgumentParser and returns the same after processing or overrides.
    """

    # If the user enabled batch mode, we disable interactive mode
    if args.batch_mode:
        args.interactive_mode = False

    if args.interactive_mode:
        override = input("Input directory? default=[{}] :".format(args.input_dir))
        if len(override):
            args.input_dir = override

        override = input("Output file? default=[{}] :".format(args.output_file))
        if len(override):
            args.output_file = override

        override = input("Mapping file? default=[{}] :".format(args.map_file))
        if len(override):
            args.map_file = override

    return(args)


def parse_arguments():
    """Get all the command line arguments and return the args from an ArgumentParser"""

    parser = argparse.ArgumentParser(
        description="A script to email students study pages for a semi-open book exam",
        epilog="Note that column count arguments start from zero."

    )

    parser.add_argument('-b', '--batch-mode',
                        action='store_true',
                        dest='batch_mode',
                        default=False,
                        help='run automatically with values given')

    parser.add_argument('--interactive-mode',
                        action='store_true',
                        dest='interactive_mode',
                        default=True,
                        help='prompt the user for details (default)')

    parser.add_argument('-i', '--input-directory',
                        dest='input_dir',
                        default='submissions',
                        help='the name of the directory containing the spreadsheets')

    parser.add_argument('-o', '--output-file',
                        dest='output_file',
                        default='output.xlsx',
                        help='the name of the file containing the output')

    parser.add_argument('-m', '--map-file',
                        dest='map_file',
                        default='mapping.cfg',
                        help='the name of the file containing the mapping')


    args = parser.parse_args()

    # Allow for any overrides from program logic or interaction with the user
    args = override_arguments(args)
    return(args)


def check_for_newrow(config_line):
    """Check if the command is a new row"""
    if config_line == "NewRow":
        return True
    else:
        return False


def check_for_block_if(config_line):
    """Check if the command is a BlockIf"""

    re_start_block = r"BlockIf:(~*)(Sheet[0-9]+-)*([A-Z]+)([0-9]+)"

    # Attempt matching
    match = re.match(re_start_block, config_line)
    if match:
        return True
    else:
        return False


def check_for_end_block_if(config_line):
    """Check if the command is a BlockIf"""

    re_end_block = r"EndBlockIf"

    # Attempt matching
    match = re.match(re_end_block, config_line)
    if match:
        return True
    else:
        return False


def select_sheet_from_workbook(workbook, sheet_name):
    """
    Attempts to select a given sheet in a workbook, and it no sheet name is given, returns the currently active sheet

    :param workbook:        The workbook in question
    :param sheet_name:      None if no name was detected, otherwise may be the sheet name with a trailing hyphen
    :return:                The selected sheet
    """

    re_sheet_match = r"^Sheet([0-9]+)-$"

    # If the sheet_name is None, then just return the active sheet
    if not sheet_name:
        return workbook.active

    # At the moment this is quite naive, and what's worse, Excel counts from 1, computers from 0
    match = re.match(re_sheet_match, sheet_name)
    if match:
        # Convert to an integer
        sheet_number = int(match.group(1))

        # Try and grab that sheet (remembering that the list is indexed from 0, not 1
        sheet_names = workbook.sheetnames
        try:
            return workbook[sheet_names[sheet_number-1]]
        except IndexError:
            print(f"Error, that sheet ({sheet_number-1}) does not exist in this spreadsheet.")
            print("Abnormal Exit.")
            sys.exit()

    # Just return the active sheet if all else fails
    return workbook.active


def process_config_line_set(config_line, source_workbook, dest_workbook, last_dest_row, match):
    """
    Handles a Set command - placing specific text in a given spreadsheet cell

    :param config_line:         The config line itself
    :param source_workbook:     The current source spreadsheet
    :param dest_workbook:       The destination spreadsheet
    :param last_dest_row:       The last row number written to
    :param match:               The match object from the regexp in process_config_line
    """

    dest_sheet_name = match.group(1)
    dest_column = match.group(2)
    dest_row = match.group(3)
    toset = match.group(4)

    # If the row is set to 0, set it to the last_dest_row
    if dest_row == '0':
        dest_row = str(last_dest_row)

    # If a destination sheet is specified, grab it, otherwise use the active sheet
    dest_sheet = select_sheet_from_workbook(dest_workbook, dest_sheet_name)

    print(f"    Set: {dest_column}{dest_row} to be {toset}")
    dest_sheet[dest_column + dest_row] = toset


def process_config_line_copy(config_line, source_workbook, dest_workbook, last_dest_row, match):
    """
    
    Handles a Copy command - copying data from one spreadsheet cell to another

    :param config_line:         The config line itself
    :param source_workbook:     The current source spreadsheet
    :param dest_workbook:       The destination spreadsheet
    :param last_dest_row:       The last row number written to
    :param match:               The match object from the regexp in process_config_line
    """

    source_sheet_name = match.group(1)
    source_column = match.group(2)
    source_row = match.group(3)

    dest_sheet_name = match.group(4)
    dest_column = match.group(5)
    dest_row = match.group(6)

    # If the dest_row is set to 0, set it to the last_dest_row
    if dest_row == '0':
        dest_row = str(last_dest_row)

    # If a destination sheet is specified, grab it, otherwise use the active sheet
    dest_sheet = select_sheet_from_workbook(dest_workbook, dest_sheet_name)

    # If a source sheet is specified, grab it, otherwise use the active sheet
    source_sheet = select_sheet_from_workbook(source_workbook, source_sheet_name)

    print(f"    Copy: {source_column}{source_row} to {dest_column}{dest_row} data: {source_sheet[source_column + source_row].value}")

    # To the final copy
    dest_sheet[dest_column + dest_row] = source_sheet[source_column + source_row].value


def stop_processing_block(config_line, source_workbook):
    """

    Determines if a block is starting, and whether we should discard config lines till it ends

    :param config_line:         The config line itself
    :param source_workbook:     The current source spreadsheet
    :return:                    True if we should stop processing, False otherwise
    """

    # The reg exp for a conditional block command
    re_start_block = r"BlockIf:(~*)(Sheet[0-9]+-)*([A-Z]+)([0-9]+)"

    # Check for a match
    match = re.match(re_start_block, config_line)

    # If we didn't get a match we're done already
    if not match:
        return False

    # Otherwise, we are down to looking at the condition look for any tilde to negate the condition
    source_not = match.group(1)
    if source_not == "~":
        source_not = False
    else:
        source_not = True

    source_sheet_name = match.group(2)
    source_column = match.group(3)
    source_row = match.group(4)

    # If a source sheet is specified, grab it, otherwise use the active sheet
    source_sheet = select_sheet_from_workbook(source_workbook, source_sheet_name)

    # Check the actual value
    data = source_sheet[source_column + source_row].value

    print(f"    BlockIf: {source_sheet_name}{source_column}{source_row} data: if {data} is {source_not}")

    if source_not:
        # We need data to be zero or empty
        if data == 0 or data is None:
            return True
        else:
            return False
    else:
        # We need data to be non zero
        if data !=0:
            return True
        else:
            return False


def process_config_line(config_line, source_workbook, dest_workbook, last_dest_row):
    """Handles processing of configuration lines, calling other helpers as required
    
    By the time this function is used the config file should have gone through two stages of pre-processing
    
    A global phase that removes comments and white space
    A local phase that has removed blocks rendered inactive by the specific source workbook

    config_line     the cleaned config line
    source_workbook the current source spreadsheet
    dest_workbook   the destingation spreadsheet
    last_dest_row   the last row number written to"""

    re_set = r"^Set:(Sheet[0-9]+-)*([A-Z]+)([0-9]+):(.*)$"
    re_copy = r"Copy:(Sheet[0-9]+-)*([A-Z]+)([0-9]+):(Sheet[0-9]+-)*([A-Z]+)([0-9]+)"

    # Is it a Set command
    match = re.match(re_set, config_line)
    if match:
        process_config_line_set(config_line, source_workbook, dest_workbook, last_dest_row, match)

    # Is it a Copy command
    match = re.match(re_copy, config_line)
    if match:
        process_config_line_copy(config_line, source_workbook, dest_workbook, last_dest_row, match)


def pre_process_config_by_whitespace(config_lines):
    """Pre process the config lines by striping leading and tailing whitespace and comments"""

    # A list for the cleaned lines
    cleaned_lines = []

    for config_line in config_lines:

        # Trim any whitespace front and end to begin with
        config_line = config_line.strip()

        # If there's any comment character trim it and to the right
        position = config_line.find('#')
        if position != -1:
            config_line = config_line[:position].strip()

        # If the resulting line isn't empty, then append it
        if config_line != "":
            cleaned_lines.append(config_line)

    return cleaned_lines


def pre_process_config_by_source(config_lines, source_workbook):
    """

    This function takes the config file and removes blocks that should not be processed based on
    the source workbook

    :param config_lines: the existing list of config lines (likely with comments removed)
    :param source_workbook: the source workbook to use to check for block removal
    :return: the config_lines that are still active
    """

    # The regular expression to show the end of the block
    re_end_block = r"EndBlockIf"

    # A list for the surviving lines
    preserved_lines = []

    # A flag for whether we are reading or not, defaults to True
    valid = True

    for config_line in config_lines:
        if stop_processing_block(config_line, source_workbook):
            # Don't copy config till later notice
            valid = False
            continue
        if re.match(re_end_block, config_line):
            # We've hit the end of a block, start reading again
            valid = True
            continue

        # We are not at the start or end of a block, so copy if and only if valid
        if valid:
            preserved_lines.append(config_line)

    print(f"  Preprocess config: {len(preserved_lines)} lines left from {len(config_lines)}")

    return preserved_lines


def process_input_directory(args):
    """
    Loops through Excel files in the input directory, mapping contents to the output file
    """

    print("Opening config...")
    # Open the config file read only and grab all the lines in it.
    try:
        config = open(args.map_file, "r")
    except OSError:
        print(f"Error opening config file {args.map_file}")
        print("Abnormal Exit.")
        sys.exit(1)
    with config:
        # Read all configuration lines, and remove comments and whitespace
        config_lines = config.read()
        config_lines = pre_process_config_by_whitespace(config_lines.splitlines())

    # Open the workbook for output
    print("Opening destination file...")
    dest_workbook = openpyxl.Workbook()
    # Activate its worksheet, the default one for now
    dest_sheet = dest_workbook.active

    # Keep track of the last written to row, for many operations that just append to the output spreadsheet
    last_dest_row = 1

    print("Searching for source files...")
    for filename in os.listdir(args.input_dir):
        filename = os.path.join(os.path.abspath(args.input_dir), filename)
        if filename.endswith("xls") or filename.endswith("xlsx"):
            print(f"  Opening {filename}...")
            # Open the source spreadsheet
            try:
                source_workbook = openpyxl.load_workbook(filename=filename, data_only=True, read_only=True)
            except OSError:
                print(f"Error opening {filename}")
                sys.exit(3)
            # A second stage pre-process of the mapping config removes any blocks conditionally for this source workbook
            preserved_lines = pre_process_config_by_source(config_lines, source_workbook)

            # Iterate through the remaining mapping configuration for this file
            for config_line in preserved_lines:
                # If we are instructed to take a new row, increment that
                if check_for_newrow(config_line):
                    last_dest_row += 1

                process_config_line(config_line, source_workbook, dest_workbook, last_dest_row)

            # Done with that source file
            source_workbook.close()

    # Save the resulting output spreadsheet
    try:
        dest_workbook.save(args.output_file)
    except OSError:
        print(f"Error saving output file {args.output_file}");
        print("Abnormal Exit.")
        sys.exit(2)


def main():
    """the main function that kicks everything else off"""

    print("excel-merger v1.1")
    args = parse_arguments()

    print("Starting excel-merger...")
    print(args)

    process_input_directory(args)

    print('Stopping excel-merger...')
    print("Successful exit.")


if __name__ == '__main__':
    main()
